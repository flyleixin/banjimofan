# -*- codeing = urf-8 -*-
# @Time :2022/5/3
# @Author:LeiXin
# @File : 班级魔方统计没填报体温.py
# @SoftWore : PyCharm
import openpyxl
import time
from selenium.webdriver import Chrome
import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
from email.mime.text import MIMEText
from email.header import Header
from smtplib import SMTP_SSL
import re
import smtplib
import socket
import os
import configparser
import datetime

# 这里是你的发信人，告诉你程序成功了没
sender =""
# 这里是QQ邮箱的授权码，不明白可以百度，或者先用着我的QQ号
pw =""







readlists = []
def read():
    wb = openpyxl.load_workbook('./xl.xlsx')
    sheet = wb['Sheet1']
    active_sheet = wb.active
    for i in range(2, 38):
        studentId = active_sheet.cell(row=i, column=1).value
        studentName = active_sheet.cell(row=i, column=2).value
        studentPost = active_sheet.cell(row=i, column=3).value
        str(studentId)
        str(studentPost)
        str(studentName)
        arr = {"studentId": studentId, "studentName": studentName, "studentPost": studentPost}
        readlists.append(arr)

lists = []
def data():
    web = Chrome()
    web.get("http://banjimofang.com/teacher/login?ref=%2Fteacher")

    web.find_element_by_xpath('/html/body/div/div/div/form/div[3]/div[1]/input').send_keys("15579460823")
    web.find_element_by_xpath('/html/body/div/div/div/form/div[3]/div[2]/input').send_keys("738836723")
    web.find_element_by_xpath('/html/body/div/div/div/form/div[3]/div[4]/button').click()
    web.find_element_by_xpath('//*[@id="mainarea"]/div[3]/div[1]/a/div/h5').click()
    web.find_element_by_xpath('/html/body/div/div[2]/div/div/div[7]/div[2]/div/div[1]/a/div[1]').click()
    web.find_element_by_xpath('/html/body/div/div[2]/div/div/div[8]/a/div[1]/i').click()
    web.switch_to.window(web.window_handles[-1])
    time.sleep(5)
    tr_list = web.find_elements_by_xpath('//*[@id="tmptablebody"]/tr')
    for tr in tr_list:
        q = tr.find_element_by_xpath("./td[3]").text
        if q == "班级":
            studentId = tr.find_element_by_xpath("./td[2]").text
            studentName = tr.find_element_by_xpath("./td[1]").text
            tem1 = tr.find_element_by_xpath("./td[4]").text
            tem2 = tr.find_element_by_xpath("./td[5]").text
            tem3 = tr.find_element_by_xpath("./td[6]").text
            tmpstudentpost = "1"
            # studentId 是str
            str(studentId)
            #print(type(studentId))


            for i in readlists:
                #print(i["studentId"])
                #print(type(i["studentId"]))
                if str(i["studentId"]) == str(studentId):
                    tmpstudentpost = i["studentPost"]
                    break
            arr = {"studentId": studentId, "studentName": studentName, "studentPost": tmpstudentpost, "tem1": tem1, "tem2": tem2, "tem3": tem3}
            lists.append(arr)
def work():
    for i in lists:
        nowtimestatus = whichtime()
        if nowtimestatus == 1:
            if str(i["tem1"]) == "0.0":
                receivers = str(i["studentPost"])
                send_msg_success(receivers)
                time.sleep(2)
                print(i["studentId"], i["studentName"], nowtimestatus, receivers)
        elif nowtimestatus == 2:
            if str(i["tem2"]) == "0.0":
                receivers = str(i["studentPost"])
                send_msg_success(receivers)
                time.sleep(2)
                print(i["studentId"], i["studentName"], nowtimestatus, receivers)
        else:
            if str(i["tem3"]) == "0.0":
                receivers = str(i["studentPost"])
                send_msg_success(receivers)
                time.sleep(2)
                print(i["studentId"], i["studentName"], nowtimestatus, receivers)


def whichtime():
    start_time1 = datetime.datetime.strptime(str(datetime.datetime.now().date()) + '0:00', '%Y-%m-%d%H:%M')
    end_time1 = datetime.datetime.strptime(str(datetime.datetime.now().date()) + '11:00', '%Y-%m-%d%H:%M')

    start_time2 = datetime.datetime.strptime(str(datetime.datetime.now().date()) + '11:00', '%Y-%m-%d%H:%M')
    end_time2 = datetime.datetime.strptime(str(datetime.datetime.now().date()) + '16:00', '%Y-%m-%d%H:%M')

    start_time3 = datetime.datetime.strptime(str(datetime.datetime.now().date()) + '17:00', '%Y-%m-%d%H:%M')
    end_time3 = datetime.datetime.strptime(str(datetime.datetime.now().date()) + '22:00', '%Y-%m-%d%H:%M')

    now_time = datetime.datetime.now()
    if start_time1 < now_time < end_time1:
        return 1
    elif start_time2 < now_time < end_time2:
        return 2
    else:
        return 3



def send_msg_success(receivers):
    server = "smtp.qq.com"
    port = 465
    machine_name = socket.gethostname()
    msg_root = MIMEMultipart('mixed')
    msg_root['From'] = Header(f'体温填报提醒')
    msg_root['Subject'] = Header(f'Message from {machine_name}', 'utf-8')
    mail_msg = f"""
	<html>

<body>
<p>您好，现在时间需要体温填报了哦！</p>
</body>
</html>"""
    msg_root.attach(MIMEText(mail_msg, 'html', 'utf-8'))
    smtp = smtplib.SMTP_SSL(server, port)
    smtp.login(sender, pw)
    smtp.sendmail(sender, receivers, msg_root.as_string())
    smtp.quit()




if __name__ == '__main__':
    try:
        read()
        data()
        work()
    except:
        pass
